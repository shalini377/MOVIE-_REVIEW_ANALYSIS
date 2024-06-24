import speech_recognition as sr
from deep_translator import GoogleTranslator
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
import pandas as pd
import pyttsx3
from gtts import gTTS
import xlsxwriter
import openpyxl
import nltk
import os

nltk.download('wordnet')
from nltk.corpus import wordnet

# Text-to-Speech initialization
engine = pyttsx3.init()
voices = engine.getProperty('voices')
engine.setProperty('rate', 150)
engine.setProperty('voice', voices[0].id)

def talk(text):
    print(text)  # Print the text to the console for better UX
    engine.say(text)
    engine.runAndWait()

def is_english_word(word):
    synsets = wordnet.synsets(word)
    return len(synsets) > 0

def write_data_to_excel(file_name, sheet_name, data):
    try:
        wb = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    sheet = wb[sheet_name] if sheet_name in wb else wb.create_sheet(sheet_name)
    for row in data:
        sheet.append(row)
    wb.save(file_name)

def update_existing_data(name, out):
    file_path = 'MOVIE.xlsx'
    df = pd.read_excel(file_path)
    df.loc[df['movie_name'] == name, out] += 1
    df.to_excel(file_path, index=False)
    print(f"{out} sentiment score for {name} has been updated.")

# Initial mode selection
talk("Mode selection. May I know what mode would you like to choose, Admin or Reviewer?")
mode = input("May I know what mode would you like to choose, Admin or Reviewer: ").lower()

# Load existing movie data
excel_file = 'MOVIE.xlsx'
df = pd.read_excel(excel_file)
column_values = df['movie_name'].tolist()

# Dictionary for sample data translations
sample_data = {
    "macha": "dude", "mamey": "dude", "semma": "bliss", "maja": "fun", "takkar": "joy",
    "sumar": "impartial", "mokka": "bad", "tharama iruku": "super quality", "tharu maru": "smash hit",
    "attu mokka": "bore", "kevalam": "awful", "vera mari": "super hit", "tharama": "quality",
    "padam": "movie", "veri a iruku": "super hit", "mokka in maruuruvam": "monotonous",
    "vera level": "out of box", "nalla illa": "not good", "palasu": "old", "puthusa onum illa": "nothing new",
    "padam": "movie", "iruku": "have"
}

if mode == "reviewer":
    print("Existing data:", column_values)
    talk("Please provide the movie name:")
    movie_name = input("Movie name: ").upper()
    positive = negative = neutral = 0

    # Microphone access and speech recognition
    recognizer = sr.Recognizer()
    with sr.Microphone() as source:
        print("Clearing the background noise...")
        recognizer.adjust_for_ambient_noise(source, duration=2)
        print("Waiting for your message...")
        recorded_audio = recognizer.listen(source)
        print("Done Recording...")

    try:
        print("Processing the message...")
        text = recognizer.recognize_google(recorded_audio, language="en-US")
        print("Your message: {}".format(text))
    except Exception as ex:
        print(ex)
        text = ""

    # Translation and dictionary mapping
    translated_text = ""
    for word in text.split():
        if is_english_word(word):
            translated_text += word + " "
        else:
            translated_text += sample_data.get(word, word) + " "

    print("Translated Text:", translated_text)

    # Sentiment Analysis
    analyser = SentimentIntensityAnalyzer()
    sentiment_scores = analyser.polarity_scores(translated_text)
    print("Sentiment Scores:", sentiment_scores)

    # Output Detection
    if sentiment_scores['compound'] > 0:
        out = "positive"
        ans = "Hey, your message sounds positive. Hurray!"
    elif sentiment_scores['compound'] == 0:
        out = "neutral"
        ans = "Hey, your message sounds neutral. Cool."
    else:
        out = "negative"
        ans = "Your message sounds sad. Don't worry!"

    print(ans)
    if out == "positive":
        positive += 1
    elif out == "negative":
        negative += 1
    elif out == "neutral":
        neutral += 1

    data = [movie_name, positive, negative, neutral]

    if movie_name not in column_values:
        write_data_to_excel('MOVIE.xlsx', 'Sheet1', [data])
    else:
        print("Already existing data")
        update_existing_data(movie_name, out)

    # System Output
    talk("Do you want me to read your review status?")
    sound = input("Do you want me to read your review status? (Y/N): ").lower()
    if sound in ["y", "yes"]:
        mytext = ans
        language = 'en'
        output = gTTS(text=mytext, lang=language, slow=False)
        output.save("output.mp3")
        os.system("start output.mp3")
    else:
        print("Thank you for your review!")
else:
    # Admin mode: Read and print the Excel file
    file_name = 'MOVIE.xlsx'
    sheet_name = 'Sheet1'
    df = pd.read_excel(file_name, sheet_name=sheet_name)
    print(df)
